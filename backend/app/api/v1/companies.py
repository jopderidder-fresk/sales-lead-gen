import hashlib
from datetime import date
from typing import Annotated, Literal

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, status
from sqlalchemy import func, nullslast, select, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.config import settings
from app.core.database import get_session
from app.core.deps import get_current_user, require_role
from app.core.logging import get_logger
from app.core.utils import LIKE_ESCAPE, escape_like, today_start_utc
from app.models.company import Company
from app.models.contact import Contact
from app.models.enrichment_job import EnrichmentJob
from app.models.icp_profile import ICPProfile
from app.models.enums import CompanyStatus, EnrichmentJobStatus, ScrapeJobStatus, SignalType
from app.models.scrape_job import ScrapeJob
from app.models.signal import Signal
from app.models.user import User
from app.schemas.company import (
    BulkDeleteRequest,
    BulkDeleteResponse,
    BulkImportResponse,
    CompanyCreate,
    CompanyDetailResponse,
    CompanyInfoResponse,
    CompanyResponse,
    CompanyUpdate,
    ImportRowError,
    PaginatedResponse,
)
from app.schemas.contact import ContactResponse
from app.schemas.deduplication import (
    DuplicateCheckRequest,
    DuplicateScanResponse,
    MergeRequest,
    SimilarCompaniesResponse,
    SimilarCompanyMatch,
)
from app.schemas.enrichment_job import EnrichmentJobResponse
from app.schemas.monitoring import MonitorTriggerResponse
from app.schemas.scrape_job import ScrapeJobResponse
from app.schemas.signal import ImportContentRequest, ImportContentResponse, SignalResponse
from app.services.deduplication import (
    find_similar_companies,
    merge_companies,
    normalize_domain,
    scan_duplicates,
    validate_public_domain,
)

logger = get_logger(__name__)

router = APIRouter(prefix="/companies", tags=["companies"])

# Allowed sort columns — whitelist to prevent injection via sort parameter
_SORT_COLUMNS = {
    "name": Company.name,
    "icp_score": Company.icp_score,
    "lead_score": Company.lead_score,
    "created_at": Company.created_at,
    "updated_at": Company.updated_at,
}


@router.post("/deduplicate", response_model=DuplicateScanResponse)
async def deduplicate_scan(
    _user: User = Depends(require_role("admin")),
    session: AsyncSession = Depends(get_session),
) -> DuplicateScanResponse:
    """Scan database for potential duplicate companies. Admin only."""
    groups = await scan_duplicates(session)
    return DuplicateScanResponse(groups=groups, total_groups=len(groups))


@router.post("/check-duplicate", response_model=SimilarCompaniesResponse)
async def check_duplicate(
    body: DuplicateCheckRequest,
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> SimilarCompaniesResponse:
    """Check if a company name/domain has potential duplicates in the database."""
    matches = await find_similar_companies(session, body.name, body.domain)
    return SimilarCompaniesResponse(matches=[SimilarCompanyMatch(**m) for m in matches])


@router.post("/merge", response_model=CompanyResponse)
async def merge(
    body: MergeRequest,
    _user: User = Depends(require_role("admin")),
    session: AsyncSession = Depends(get_session),
) -> Company:
    """Merge a duplicate company into a primary company. Admin only.

    Reassigns all contacts, signals, and scrape jobs from the duplicate to the primary,
    then deletes the duplicate.
    """
    try:
        primary = await merge_companies(session, body.primary_id, body.duplicate_id)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(exc),
        ) from None
    logger.info(
        "Companies merged via API",
        primary_id=body.primary_id,
        duplicate_id=body.duplicate_id,
    )
    return primary


# Dutch Excel header → resolved field name.  Keys prefixed with ``_bd.``
# are collected into the ``bedrijfsdata`` JSONB column; all others map
# directly to Company model columns.
_BEDRIJFSDATA_COLUMN_MAP: dict[str, str] = {
    # Direct field mappings (Dutch header → English DB column)
    "bedrijfsnaam": "name",
    "domein": "domain",
    "kvk": "kvk_number",
    "adres": "address",
    "postcode": "postal_code",
    "plaats": "city",
    "provincie": "province",
    "land": "country",
    "telefoon": "phone",
    "telefoon_volledig": "phone",  # preferred over telefoon
    "website": "website_url",
    "emailadres": "email",
    "oprichtingsjaar": "founded_year",
    "medewerkers": "employee_count",
    "medewerkers_range": "size",
    "branches_kvk": "industry",
    "linkedin_link": "linkedin_url",
    "facebook_link": "facebook_url",
    "twitter_link": "twitter_url",
    "organisatietype": "organization_type",
    # Fields routed into the bedrijfsdata JSONB
    "bedrijfsdata_id": "_bd.bedrijfsdata_id",
    "gemeente": "_bd.gemeente",
    "vestigingstype": "_bd.vestigingstype",
    "btwnummer": "_bd.btwnummer",
    "sbi_codes": "_bd.sbi_codes",
    "coordinaten": "_bd.coordinaten",
    "cms": "_bd.cms",
    "website_analytics": "_bd.website_analytics",
    "cdn": "_bd.cdn",
    "advertentienetwerken": "_bd.advertentienetwerken",
    "caching_server": "_bd.caching_server",
    "webshop": "_bd.webshop",
    "emailprovider": "_bd.emailprovider",
    "apps": "_bd.apps",
    "bedrijfsprofiel": "_bd.bedrijfsprofiel",
    "youtube_link": "_bd.youtube_link",
    "instagram_link": "_bd.instagram_link",
    "pinterest_link": "_bd.pinterest_link",
}


def _cell_str(row: tuple, col_map: dict[str, int], key: str) -> str | None:
    """Return a stripped string from *row* at the column mapped to *key*, or ``None``."""
    if key not in col_map:
        return None
    idx = col_map[key]
    if idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    # Excel stores numbers as floats; strip ".0" for values that are integers
    # (e.g. KVK numbers: 16045656.0 → "16045656").
    if isinstance(val, float) and val == int(val):
        val = int(val)
    text = str(val).strip()
    return text or None


def _safe_int(val: object) -> int | None:
    if val is None:
        return None
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return None


@router.post("/import", response_model=BulkImportResponse)
async def import_companies(
    file: UploadFile,
    _user: User = Depends(require_role("admin", "user")),
    session: AsyncSession = Depends(get_session),
) -> BulkImportResponse:
    """Import companies from an Excel (.xlsx) file.

    Accepts both English headers (name, domain, industry, size, location) and
    Dutch Bedrijfsdata headers (bedrijfsnaam, domein, branches_kvk, etc.).
    First row must be the header.  Maximum 500 data rows.
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Only .xlsx and .xls files are supported",
        )

    # Limit file size to 5 MB
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="File too large. Maximum 5 MB.",
        )

    import io

    from openpyxl import load_workbook

    try:
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Cannot read Excel file: {exc}",
        ) from exc

    ws = wb.active
    if ws is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Excel file has no active worksheet",
        )

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="File must have a header row and at least one data row",
        )

    # Parse header (case-insensitive) and resolve Dutch names to English via
    # the column map.  ``telefoon_volledig`` intentionally appears *after*
    # ``telefoon`` in the Excel so it overwrites the shorter variant.
    raw_header = [str(cell).strip().lower() if cell else "" for cell in rows[0]]
    resolved: list[str] = []
    for h in raw_header:
        resolved.append(_BEDRIJFSDATA_COLUMN_MAP.get(h, h))

    # Build col_map; later occurrences override earlier ones so that e.g.
    # ``telefoon_volledig`` (→ "phone") wins over ``telefoon`` (→ "phone").
    col_map: dict[str, int] = {}
    for idx, field in enumerate(resolved):
        col_map[field] = idx

    has_name = "name" in col_map
    has_domain = "domain" in col_map or "website_url" in col_map
    if not has_name or not has_domain:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Header must contain 'name'/'bedrijfsnaam' and 'domain'/'domein'/'website' columns",
        )

    data_rows = rows[1:501]  # Max 500 rows

    # Pre-load existing (name, domain) pairs for dedup
    existing = await session.execute(
        select(Company.name, Company.domain).where(Company.status != CompanyStatus.ARCHIVED)
    )
    existing_pairs: set[tuple[str, str]] = {
        (r.name, r.domain) for r in existing.all() if r.domain
    }

    imported = 0
    skipped = 0
    errors: list[ImportRowError] = []

    for row_idx, row in enumerate(data_rows, start=2):
        try:
            name = _cell_str(row, col_map, "name") or ""
            # Domain: prefer dedicated domain column, fall back to website_url
            domain_raw = _cell_str(row, col_map, "domain") or _cell_str(row, col_map, "website_url") or ""

            if not name or not domain_raw:
                errors.append(ImportRowError(row=row_idx, error="Name and domain are required"))
                continue

            domain_error = validate_public_domain(domain_raw)
            if domain_error:
                errors.append(ImportRowError(row=row_idx, error=domain_error))
                continue

            normalized = normalize_domain(domain_raw)
            if (name, normalized) in existing_pairs:
                skipped += 1
                continue

            # Build location from city + province (matching discovery.py pattern)
            city_val = _cell_str(row, col_map, "city")
            province_val = _cell_str(row, col_map, "province")
            location_parts = [p for p in [city_val, province_val] if p]
            location = ", ".join(location_parts) if location_parts else _cell_str(row, col_map, "location")

            # Collect bedrijfsdata JSONB fields
            bd_data: dict[str, str] = {}
            for field_name, idx in col_map.items():
                if field_name.startswith("_bd.") and idx < len(row):
                    val = row[idx]
                    if val is not None:
                        text = str(val).strip()
                        if text:
                            bd_data[field_name[4:]] = text

            company = Company(
                name=name,
                domain=normalized,
                industry=_cell_str(row, col_map, "industry"),
                size=_cell_str(row, col_map, "size"),
                location=location,
                kvk_number=_cell_str(row, col_map, "kvk_number"),
                phone=_cell_str(row, col_map, "phone"),
                email=_cell_str(row, col_map, "email"),
                website_url=_cell_str(row, col_map, "website_url"),
                address=_cell_str(row, col_map, "address"),
                postal_code=_cell_str(row, col_map, "postal_code"),
                city=city_val,
                province=province_val,
                country=_cell_str(row, col_map, "country"),
                founded_year=_safe_int(_cell_str(row, col_map, "founded_year")),
                employee_count=_safe_int(_cell_str(row, col_map, "employee_count")),
                organization_type=_cell_str(row, col_map, "organization_type"),
                linkedin_url=_cell_str(row, col_map, "linkedin_url"),
                facebook_url=_cell_str(row, col_map, "facebook_url"),
                twitter_url=_cell_str(row, col_map, "twitter_url"),
                bedrijfsdata=bd_data if bd_data else None,
                status=CompanyStatus.DISCOVERED,
            )
            async with session.begin_nested():
                session.add(company)
                await session.flush()
            imported += 1
            existing_pairs.add((name, normalized))

        except Exception as exc:
            errors.append(ImportRowError(row=row_idx, error=str(exc)))

    await session.commit()
    wb.close()

    logger.info("companies.import", imported=imported, skipped=skipped, errors=len(errors))
    return BulkImportResponse(imported=imported, skipped=skipped, errors=errors)


@router.post("/bulk-delete", response_model=BulkDeleteResponse)
async def bulk_delete_companies(
    body: BulkDeleteRequest,
    _user: User = Depends(require_role("admin", "user")),
    session: AsyncSession = Depends(get_session),
) -> BulkDeleteResponse:
    """Archive multiple companies at once."""
    cursor = await session.execute(
        update(Company)
        .where(Company.id.in_(body.ids), Company.status != CompanyStatus.ARCHIVED)
        .values(status=CompanyStatus.ARCHIVED)
    )
    await session.commit()

    count = cursor.rowcount  # type: ignore[attr-defined]
    logger.info("companies.bulk_delete", archived=count, requested=len(body.ids))
    return BulkDeleteResponse(archived=count)


@router.get("", response_model=PaginatedResponse[CompanyResponse])
async def list_companies(
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=25, ge=1, le=100),
    status_filter: CompanyStatus | None = Query(default=None, alias="status"),
    industry: str | None = Query(default=None, max_length=255),
    min_score: float | None = Query(default=None, ge=0),
    min_lead_score: float | None = Query(default=None, ge=0),
    monitor: bool | None = Query(default=None),
    search: str | None = Query(default=None, max_length=255),
    added_after: date | None = Query(default=None),
    added_before: date | None = Query(default=None),
    sort: str = Query(default="created_at"),
    order: Literal["asc", "desc"] = Query(default="desc"),
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> PaginatedResponse[CompanyResponse]:
    """List companies with pagination, filtering, sorting, and search."""
    query = select(Company)

    # When a status filter is given, apply it directly (including 'archived').
    # Otherwise exclude archived by default so the default listing stays clean.
    if status_filter is not None:
        query = query.where(Company.status == status_filter)
    else:
        query = query.where(Company.status != CompanyStatus.ARCHIVED)

    # Filtering
    if industry is not None:
        pattern = f"%{escape_like(industry)}%"
        query = query.where(Company.industry.ilike(pattern, escape=LIKE_ESCAPE))
    if min_score is not None:
        query = query.where(Company.icp_score >= min_score)
    if min_lead_score is not None:
        query = query.where(Company.lead_score >= min_lead_score)
    if monitor is not None:
        query = query.where(Company.monitor == monitor)

    # Date range filter
    if added_after is not None:
        query = query.where(func.date(Company.created_at) >= added_after)
    if added_before is not None:
        query = query.where(func.date(Company.created_at) <= added_before)

    # Search (ILIKE on name and domain)
    if search is not None:
        escaped = escape_like(search)
        pattern = f"%{escaped}%"
        query = query.where(
            Company.name.ilike(pattern, escape=LIKE_ESCAPE)
            | Company.domain.ilike(pattern, escape=LIKE_ESCAPE)
        )

    # Total count before pagination
    count_query = select(func.count()).select_from(query.subquery())
    total = (await session.execute(count_query)).scalar_one()

    # Sorting — only allow whitelisted columns
    sort_column = _SORT_COLUMNS.get(sort)
    if sort_column is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid sort field. Allowed: {', '.join(_SORT_COLUMNS)}",
        )
    sort_clause = sort_column.desc() if order == "desc" else sort_column.asc()
    query = query.order_by(nullslast(sort_clause), Company.id.asc())

    # Pagination
    query = query.offset(offset).limit(limit).options(selectinload(Company.crm_integration))

    result = await session.execute(query)
    companies = result.scalars().all()

    return PaginatedResponse(
        items=[CompanyResponse.model_validate(c) for c in companies],
        total=total,
        offset=offset,
        limit=limit,
    )


@router.get("/{company_id}", response_model=CompanyDetailResponse)
async def get_company(
    company_id: int,
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> CompanyDetailResponse:
    """Get company detail including contacts count, signals count, and latest signal."""
    result = await session.execute(
        select(Company)
        .where(Company.id == company_id)
        .options(selectinload(Company.crm_integration))
    )
    company = result.scalar_one_or_none()

    if company is None or company.status == CompanyStatus.ARCHIVED:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Company not found",
        )

    # Separate count queries to avoid cross-join inflating counts
    contacts_count = (
        await session.execute(select(func.count()).where(Contact.company_id == company_id))
    ).scalar_one()

    signals_result = await session.execute(
        select(
            func.count().label("signals_count"),
            func.max(Signal.created_at).label("latest_signal_at"),
        ).where(Signal.company_id == company_id)
    )
    signals_row = signals_result.one()

    return CompanyDetailResponse(
        **CompanyResponse.model_validate(company).model_dump(),
        contacts_count=contacts_count,
        signals_count=signals_row.signals_count,
        latest_signal_at=signals_row.latest_signal_at,
        company_info=CompanyInfoResponse(**company.company_info) if company.company_info else None,
    )


@router.post("", response_model=CompanyResponse, status_code=status.HTTP_201_CREATED)
async def create_company(
    body: CompanyCreate,
    _user: User = Depends(require_role("admin", "user")),
    session: AsyncSession = Depends(get_session),
) -> Company:
    """Create a new company (used by discovery engine and manual entry)."""
    normalized = normalize_domain(body.domain)

    # SSRF prevention: reject IP addresses, localhost, and internal hostnames
    domain_error = validate_public_domain(body.domain)
    if domain_error:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=domain_error,
        )

    # Optimistic check — catches most duplicates cheaply before hitting the DB constraint.
    existing = await session.execute(
        select(Company).where(Company.name == body.name, Company.domain == normalized)
    )
    if existing.scalar_one_or_none() is not None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="A company with this name and domain already exists",
        )

    company = Company(**{**body.model_dump(), "domain": normalized})
    session.add(company)
    try:
        await session.commit()
    except IntegrityError as error:
        await session.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="A company with this name and domain already exists",
        ) from error
    await session.refresh(company)

    logger.info("Company created", company_id=company.id, domain=company.domain)
    return company


@router.put("/{company_id}", response_model=CompanyResponse)
async def update_company(
    company_id: int,
    body: CompanyUpdate,
    _user: User = Depends(require_role("admin", "user")),
    session: AsyncSession = Depends(get_session),
) -> Company:
    """Update company fields."""
    result = await session.execute(
        select(Company)
        .where(Company.id == company_id)
        .options(selectinload(Company.crm_integration))
    )
    company = result.scalar_one_or_none()

    if company is None or company.status == CompanyStatus.ARCHIVED:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Company not found",
        )

    # Optimistic (name, domain) uniqueness check before hitting the DB constraint.
    update_data = body.model_dump(exclude_unset=True)
    if "domain" in update_data:
        domain_error = validate_public_domain(update_data["domain"])
        if domain_error:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=domain_error,
            )
        update_data["domain"] = normalize_domain(update_data["domain"])
    new_name = update_data.get("name", company.name)
    new_domain = update_data.get("domain", company.domain)
    if new_name != company.name or new_domain != company.domain:
        existing = await session.execute(
            select(Company).where(
                Company.name == new_name,
                Company.domain == new_domain,
                Company.id != company_id,
            )
        )
        if existing.scalar_one_or_none() is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="A company with this name and domain already exists",
            )

    # When the user explicitly toggles monitor, pin it so auto-logic won't override
    if "monitor" in update_data and "monitor_pinned" not in update_data:
        update_data["monitor_pinned"] = True

    for field, value in update_data.items():
        setattr(company, field, value)

    try:
        await session.commit()
    except IntegrityError as error:
        await session.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="A company with this domain already exists",
        ) from error
    await session.refresh(company, attribute_names=["crm_integration"])

    logger.info("Company updated", company_id=company.id)
    return company


@router.delete("/{company_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_company(
    company_id: int,
    _user: User = Depends(require_role("admin", "user")),
    session: AsyncSession = Depends(get_session),
) -> None:
    """Soft delete a company by setting status to archived."""
    result = await session.execute(select(Company).where(Company.id == company_id))
    company = result.scalar_one_or_none()

    if company is None or company.status == CompanyStatus.ARCHIVED:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Company not found",
        )

    company.status = CompanyStatus.ARCHIVED
    await session.commit()

    logger.info("Company archived", company_id=company.id, domain=company.domain)


@router.post(
    "/{company_id}/monitor",
    response_model=MonitorTriggerResponse,
    status_code=status.HTTP_202_ACCEPTED,
    summary="Trigger signal monitoring for a company",
)
async def trigger_monitor(
    company_id: int,
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> MonitorTriggerResponse:
    """Manually trigger signal monitoring for a specific company.

    Dispatches a Celery task that scrapes the company's configured pages,
    detects content changes, and creates Signal records for the LLM pipeline.
    """
    await _get_company_or_404(session, company_id)

    from app.tasks.monitoring import monitor_company_task

    try:
        task = monitor_company_task.delay(company_id)
    except Exception as exc:
        logger.error("monitor.trigger_failed", company_id=company_id, error=str(exc))
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Failed to dispatch monitoring task. Is the task queue running?",
        ) from exc

    logger.info("monitor.triggered", company_id=company_id, task_id=task.id)
    return MonitorTriggerResponse(
        task_id=task.id,
        company_id=company_id,
        message="Monitoring task dispatched",
    )


@router.post(
    "/{company_id}/linkedin-scrape",
    response_model=MonitorTriggerResponse,
    status_code=status.HTTP_202_ACCEPTED,
    summary="Trigger LinkedIn scrape for a company via Apify",
    tags=["enrichment"],
)
async def trigger_linkedin_scrape(
    company_id: int,
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> MonitorTriggerResponse:
    """Scrape a company's LinkedIn page for posts and company data.

    Requires the company to have a ``linkedin_url`` set. Creates Signal
    records that the LLM pipeline picks up for classification and scoring.
    """
    company = await _get_company_or_404(session, company_id)

    if not company.linkedin_url:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Company has no LinkedIn URL set. Scrape the website first or add it manually.",
        )

    from app.tasks.linkedin import scrape_company_linkedin

    try:
        task = scrape_company_linkedin.delay(company_id)
    except Exception as exc:
        logger.error("linkedin_scrape.trigger_failed", company_id=company_id, error=str(exc))
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Failed to dispatch LinkedIn scrape task. Is the task queue running?",
        ) from exc

    logger.info("linkedin_scrape.triggered", company_id=company_id, task_id=task.id)
    return MonitorTriggerResponse(
        task_id=task.id,
        company_id=company_id,
        message="LinkedIn scrape task dispatched",
    )


# ── Sub-resource helpers ────────────────────────────────────────────────────


async def _require_active_icp(session: AsyncSession) -> None:
    """Raise 422 if no ICP profile is active."""
    result = await session.execute(
        select(ICPProfile.id).where(ICPProfile.is_active.is_(True)).limit(1)
    )
    if result.scalar_one_or_none() is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail="No active ICP profile. Please activate an ICP profile first.",
        )


async def _get_company_or_404(session: AsyncSession, company_id: int) -> Company:
    """Fetch a non-archived company or raise 404."""
    result = await session.execute(select(Company).where(Company.id == company_id))
    company = result.scalar_one_or_none()
    if company is None or company.status == CompanyStatus.ARCHIVED:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Company not found",
        )
    return company


# ── Sub-resource endpoints ──────────────────────────────────────────────────


@router.get("/{company_id}/contacts", response_model=PaginatedResponse[ContactResponse])
async def list_company_contacts(
    company_id: int,
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=25, ge=1, le=100),
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> PaginatedResponse[ContactResponse]:
    """List contacts for a company, ordered by creation date descending."""
    await _get_company_or_404(session, company_id)

    query = select(Contact).where(Contact.company_id == company_id)
    total = (await session.execute(select(func.count()).select_from(query.subquery()))).scalar_one()

    query = query.order_by(Contact.created_at.desc()).offset(offset).limit(limit)
    contacts = (await session.execute(query)).scalars().all()

    return PaginatedResponse(
        items=[ContactResponse.model_validate(c) for c in contacts],
        total=total,
        offset=offset,
        limit=limit,
    )


@router.get("/{company_id}/signals", response_model=PaginatedResponse[SignalResponse])
async def list_company_signals(
    company_id: int,
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=25, ge=1, le=100),
    signal_type: Annotated[list[SignalType], Query()] = [],  # noqa: B006
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> PaginatedResponse[SignalResponse]:
    """List signals for a company, ordered by creation date descending."""
    await _get_company_or_404(session, company_id)

    query = select(Signal).where(Signal.company_id == company_id)
    if signal_type:
        query = query.where(Signal.signal_type.in_(signal_type))

    total = (await session.execute(select(func.count()).select_from(query.subquery()))).scalar_one()

    query = query.order_by(Signal.created_at.desc()).offset(offset).limit(limit)
    signals = (await session.execute(query)).scalars().all()

    return PaginatedResponse(
        items=[SignalResponse.model_validate(s) for s in signals],
        total=total,
        offset=offset,
        limit=limit,
    )


@router.get("/{company_id}/scrape-jobs", response_model=PaginatedResponse[ScrapeJobResponse])
async def list_company_scrape_jobs(
    company_id: int,
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=25, ge=1, le=100),
    job_status: ScrapeJobStatus | None = Query(default=None, alias="status"),
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> PaginatedResponse[ScrapeJobResponse]:
    """List scrape jobs for a company, ordered by creation date descending."""
    await _get_company_or_404(session, company_id)

    query = select(ScrapeJob).where(ScrapeJob.company_id == company_id)
    if job_status is not None:
        query = query.where(ScrapeJob.status == job_status)

    total = (await session.execute(select(func.count()).select_from(query.subquery()))).scalar_one()

    query = query.order_by(ScrapeJob.created_at.desc()).offset(offset).limit(limit)
    jobs = (await session.execute(query)).scalars().all()

    return PaginatedResponse(
        items=[ScrapeJobResponse.model_validate(j) for j in jobs],
        total=total,
        offset=offset,
        limit=limit,
    )


@router.get(
    "/{company_id}/enrichment-jobs", response_model=PaginatedResponse[EnrichmentJobResponse]
)
async def list_company_enrichment_jobs(
    company_id: int,
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=25, ge=1, le=100),
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> PaginatedResponse[EnrichmentJobResponse]:
    """List enrichment jobs for a company, ordered by creation date descending."""
    await _get_company_or_404(session, company_id)

    query = select(EnrichmentJob).where(EnrichmentJob.company_id == company_id)
    total = (await session.execute(select(func.count()).select_from(query.subquery()))).scalar_one()

    query = query.order_by(EnrichmentJob.created_at.desc()).offset(offset).limit(limit)
    jobs = (await session.execute(query)).scalars().all()

    return PaginatedResponse(
        items=[EnrichmentJobResponse.model_validate(j) for j in jobs],
        total=total,
        offset=offset,
        limit=limit,
    )


# ── Enrichment & Scrape triggers ──────────────────────────────────────────


@router.post(
    "/{company_id}/enrich",
    response_model=MonitorTriggerResponse,
    status_code=status.HTTP_202_ACCEPTED,
    summary="Trigger LLM enrichment for a company",
    tags=["enrichment"],
)
async def trigger_enrich(
    company_id: int,
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> MonitorTriggerResponse:
    """Dispatch LLM enrichment: generate company profile + analyze signals.

    Uses existing scraped content from Signal records. Never scrapes.
    """
    await _require_active_icp(session)
    await _get_company_or_404(session, company_id)

    # Check daily enrichment limit
    enrichments_today = (
        await session.execute(
            select(func.count())
            .select_from(EnrichmentJob)
            .where(
                EnrichmentJob.created_at >= today_start_utc(),
            )
        )
    ).scalar_one()
    if enrichments_today >= settings.max_enrichments_per_day:
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail=f"Daily enrichment limit reached ({settings.max_enrichments_per_day}/day). "
            "Adjust in Settings > Usage Limits.",
        )

    # Create an EnrichmentJob record so status is trackable in the UI
    job = EnrichmentJob(company_id=company_id, status=EnrichmentJobStatus.PENDING)
    session.add(job)
    await session.commit()
    await session.refresh(job)

    from app.tasks.enrichment import enrich_company

    try:
        task = enrich_company.delay(company_id, job.id)
    except Exception as exc:
        logger.error("enrich.trigger_failed", company_id=company_id, error=str(exc))
        job.status = EnrichmentJobStatus.FAILED
        job.error_message = "Failed to dispatch task"
        await session.commit()
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Failed to dispatch enrichment task. Is the task queue running?",
        ) from exc

    logger.info("enrich.triggered", company_id=company_id, task_id=task.id, job_id=job.id)
    return MonitorTriggerResponse(
        task_id=task.id,
        company_id=company_id,
        message="Enrichment task dispatched",
    )


@router.post(
    "/{company_id}/contacts",
    response_model=MonitorTriggerResponse,
    status_code=status.HTTP_202_ACCEPTED,
    summary="Trigger contact finding for a company",
    tags=["enrichment"],
)
async def trigger_contacts(
    company_id: int,
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> MonitorTriggerResponse:
    """Dispatch contact finding via waterfall: Hunter.io, ScrapIn, LLM.

    Never scrapes websites. Uses already-scraped Signal data for LLM fallback.
    """
    await _require_active_icp(session)
    await _get_company_or_404(session, company_id)

    # Create an EnrichmentJob record so status is trackable in the UI
    job = EnrichmentJob(company_id=company_id, status=EnrichmentJobStatus.PENDING)
    session.add(job)
    await session.commit()
    await session.refresh(job)

    from app.tasks.contacts import find_company_contacts

    try:
        task = find_company_contacts.delay(company_id, job.id)
    except Exception as exc:
        logger.error("contacts.trigger_failed", company_id=company_id, error=str(exc))
        job.status = EnrichmentJobStatus.FAILED
        job.error_message = "Failed to dispatch task"
        await session.commit()
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Failed to dispatch contacts task. Is the task queue running?",
        ) from exc

    logger.info("contacts.triggered", company_id=company_id, task_id=task.id, job_id=job.id)
    return MonitorTriggerResponse(
        task_id=task.id,
        company_id=company_id,
        message="Contact finding task dispatched",
    )


@router.post(
    "/{company_id}/scrape",
    response_model=MonitorTriggerResponse,
    status_code=status.HTTP_202_ACCEPTED,
    summary="Trigger a Firecrawl scrape for a company",
    tags=["enrichment"],
)
async def trigger_scrape(
    company_id: int,
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> MonitorTriggerResponse:
    """Create a scrape job record and dispatch a Firecrawl crawl task
    that scrapes the company's domain (about, team, careers, blog, news pages).
    """
    await _require_active_icp(session)
    company = await _get_company_or_404(session, company_id)

    # Check daily scrape limit
    scrapes_today = (
        await session.execute(
            select(func.count())
            .select_from(ScrapeJob)
            .where(
                ScrapeJob.created_at >= today_start_utc(),
            )
        )
    ).scalar_one()
    if scrapes_today >= settings.max_scrapes_per_day:
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail=f"Daily scrape limit reached ({settings.max_scrapes_per_day}/day). "
            "Adjust in Settings > Usage Limits.",
        )

    # Reject if this company was already scraped today
    company_scrapes_today = (
        await session.execute(
            select(func.count())
            .select_from(ScrapeJob)
            .where(
                ScrapeJob.company_id == company_id,
                ScrapeJob.created_at >= today_start_utc(),
            )
        )
    ).scalar_one()
    if company_scrapes_today > 0:
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail="This company has already been scraped today. Try again tomorrow.",
        )

    # Reject if a scrape job is already pending or running
    active_job = (
        await session.execute(
            select(ScrapeJob)
            .where(
                ScrapeJob.company_id == company_id,
                ScrapeJob.status.in_([ScrapeJobStatus.PENDING, ScrapeJobStatus.RUNNING]),
            )
            .limit(1)
        )
    ).scalar_one_or_none()
    if active_job:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="A scrape job is already running for this company.",
        )

    # Create a ScrapeJob record
    job = ScrapeJob(
        company_id=company_id,
        target_url=f"https://{company.domain}",
        status=ScrapeJobStatus.PENDING,
    )
    session.add(job)
    await session.commit()
    await session.refresh(job)

    from app.tasks.scraping import trigger_company_scrape

    try:
        trigger_company_scrape.delay(company_id, job.id)
    except Exception as exc:
        logger.error("scrape.dispatch_failed", company_id=company_id, error=str(exc))
        job.status = ScrapeJobStatus.FAILED
        job.error_message = "Failed to dispatch scrape task"
        await session.commit()
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Failed to dispatch scrape task. Is the task queue running?",
        ) from exc

    logger.info("scrape.dispatched", company_id=company_id, job_id=job.id)
    return MonitorTriggerResponse(
        task_id=str(job.id),
        company_id=company_id,
        message="Scrape task dispatched",
    )


@router.post(
    "/{company_id}/pipeline",
    response_model=MonitorTriggerResponse,
    status_code=status.HTTP_202_ACCEPTED,
    summary="Run full pipeline: scrape → linkedin → enrich → contacts",
    tags=["enrichment"],
)
async def trigger_pipeline(
    company_id: int,
    _user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> MonitorTriggerResponse:
    """Run the full pipeline for a company: scrape → linkedin → enrich → contacts.

    Uses Celery chain to run the four tasks in sequence.
    """
    await _require_active_icp(session)
    company = await _get_company_or_404(session, company_id)

    # Create job records for tracking
    scrape_job = ScrapeJob(
        company_id=company_id,
        target_url=f"https://{company.domain}",
        status=ScrapeJobStatus.PENDING,
    )
    enrich_job = EnrichmentJob(company_id=company_id, status=EnrichmentJobStatus.PENDING)
    contacts_job = EnrichmentJob(company_id=company_id, status=EnrichmentJobStatus.PENDING)
    session.add_all([scrape_job, enrich_job, contacts_job])
    await session.commit()
    await session.refresh(scrape_job)
    await session.refresh(enrich_job)
    await session.refresh(contacts_job)

    from celery import chain

    from app.tasks.contacts import find_company_contacts
    from app.tasks.enrichment import enrich_company
    from app.tasks.linkedin import scrape_company_linkedin_safe
    from app.tasks.scraping import trigger_company_scrape

    try:
        pipeline = chain(
            trigger_company_scrape.si(company_id, scrape_job.id),
            scrape_company_linkedin_safe.si(company_id),
            enrich_company.si(company_id, enrich_job.id),
            find_company_contacts.si(company_id, contacts_job.id),
        )
        result = pipeline.apply_async()
    except Exception as exc:
        logger.error("pipeline.trigger_failed", company_id=company_id, error=str(exc))
        scrape_job.status = ScrapeJobStatus.FAILED
        scrape_job.error_message = "Failed to dispatch pipeline"
        enrich_job.status = EnrichmentJobStatus.FAILED
        enrich_job.error_message = "Failed to dispatch pipeline"
        contacts_job.status = EnrichmentJobStatus.FAILED
        contacts_job.error_message = "Failed to dispatch pipeline"
        await session.commit()
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Failed to dispatch pipeline. Is the task queue running?",
        ) from exc

    logger.info(
        "pipeline.triggered",
        company_id=company_id,
        task_id=result.id,
        scrape_job_id=scrape_job.id,
        enrich_job_id=enrich_job.id,
        contacts_job_id=contacts_job.id,
    )
    return MonitorTriggerResponse(
        task_id=result.id,
        company_id=company_id,
        message="Pipeline dispatched: scrape → linkedin → enrich → contacts",
    )


@router.post(
    "/{company_id}/import-content",
    response_model=ImportContentResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Import scraped content as signals for a company",
    tags=["enrichment"],
)
async def import_content(
    company_id: int,
    body: list[ImportContentRequest],
    _user: User = Depends(require_role("admin", "user")),
    session: AsyncSession = Depends(get_session),
) -> ImportContentResponse:
    """Import externally scraped content into the database as Signal records.

    Accepts a list of scraped pages (source URL + markdown) and creates
    unprocessed signals that the LLM pipeline will pick up automatically.
    Duplicate content (same hash for the same company) is skipped.
    """
    company = await _get_company_or_404(session, company_id)

    # Cap the number of pages to prevent abuse / DoS
    _MAX_IMPORT_ITEMS = 50
    if len(body) > _MAX_IMPORT_ITEMS:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Too many items: {len(body)}. Maximum is {_MAX_IMPORT_ITEMS} per request.",
        )

    existing_hashes_result = await session.execute(
        select(Signal.raw_content_hash).where(
            Signal.company_id == company_id,
            Signal.raw_content_hash.isnot(None),
        )
    )
    existing_hashes = set(existing_hashes_result.scalars().all())

    signal_ids: list[int] = []
    for item in body:
        if len(item.markdown.strip()) < 50:
            continue
        normalized = " ".join(item.markdown.split())
        content_hash = hashlib.sha256(normalized.encode()).hexdigest()
        if content_hash in existing_hashes:
            continue
        existing_hashes.add(content_hash)

        signal = Signal(
            company_id=company_id,
            source_url=item.source_url,
            signal_type=SignalType.NO_SIGNAL,
            raw_markdown=item.markdown,
            raw_content_hash=content_hash,
            is_processed=False,
        )
        session.add(signal)
        await session.flush()
        signal_ids.append(signal.id)

    await session.commit()

    if signal_ids:
        from app.services.intelligence import analyze_signal_ids_inline

        try:
            await analyze_signal_ids_inline(signal_ids)
        except Exception:
            logger.exception(
                "import.intelligence_failed",
                company_id=company_id,
                signal_count=len(signal_ids),
            )

        from app.tasks.lead_scoring import recalculate_company_score

        try:
            recalculate_company_score.delay(company_id)
        except Exception:
            logger.warning("import.score_recalc_failed", company_id=company_id)

    created = len(signal_ids)
    logger.info(
        "import.content",
        company_id=company_id,
        domain=company.domain,
        signals_created=created,
        pages_submitted=len(body),
    )
    return ImportContentResponse(
        signals_created=created,
        company_id=company_id,
        message=f"Imported {created} signal(s) from {len(body)} page(s)",
    )
